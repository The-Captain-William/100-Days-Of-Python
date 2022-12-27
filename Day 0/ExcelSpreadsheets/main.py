import openpyxl

# wb = openpyxl.Workbook()  # workbook in memory
wb = openpyxl.load_workbook("transactions.xlsx")  # open workbook
print(wb.sheetnames)  # won't pull up automatically

sheet = wb["Sheet1"]  # case-sensitive. Creating a sheet object for actual sheet.

# wb.create_sheet("Sheet2")  # also has index parameter (sheet name,index). No brackets
# wb.remove_sheet will remove sheet

cell = sheet["a1"]  # looks for cell ID in Excel.
#  print(sheet.max_row) total # of columns and rows
#  print(sheet.max_column)

for row in range(1, sheet.max_row + 1): # every row from 1 to max
    for column in range(1, sheet.max_column + 1):  # same but columns
        cell = sheet.cell(row, column)
        print(cell.value)       # will print contents

sheet.append([1, 2, 3])

wb.save("transactions2.xlsx")
